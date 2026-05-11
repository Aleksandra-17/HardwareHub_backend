"""Бизнес-логика роутера reports."""

import csv
import io
import uuid
from decimal import Decimal

from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.routers.components.models import Component, ComputerComponent
from src.routers.device_types.models import DeviceType
from src.routers.devices.models import Device
from src.routers.licenses.models import License
from src.routers.locations.models import Location
from src.routers.people.models import Person
from src.routers.reports.schemas import (
    ExportRequest,
    InventoryDeviceItem,
    InventoryRequest,
    InventoryResponse,
)
from src.routers.workstations.models import Workstation

STATUS_LABELS = {
    "in_use": "в использовании",
    "repair": "в ремонте",
    "scrapped": "списано",
    "archived": "архив",
}

CATEGORY_LABELS = {
    "computing": "Вычислительная техника",
    "office": "Оргтехника",
    "network": "Сетевое оборудование",
    "peripheral": "Периферия",
    "other": "Прочее",
}

COMPONENT_TYPE_LABELS = {
    "cpu": "Процессор",
    "motherboard": "Материнская плата",
    "ram": "ОЗУ",
    "storage": "Накопитель",
    "psu": "Блок питания",
    "gpu": "Видеокарта",
    "case": "Корпус",
    "cooler": "Охлаждение",
}


async def export_devices(
    session: AsyncSession,
    req: ExportRequest,
) -> tuple[bytes, str, str]:
    """Экспорт устройств в CSV или Excel. Возвращает (content, content_type, filename)."""
    stmt = (
        select(
            Device,
            DeviceType.name.label("type_name"),
            DeviceType.category.label("type_category"),
            Location.name.label("location_name"),
            Person.full_name.label("person_name"),
            Workstation.seat_code.label("workstation_seat"),
        )
        .outerjoin(DeviceType, Device.device_type_id == DeviceType.id)
        .outerjoin(Location, Device.location_id == Location.id)
        .outerjoin(Person, Device.person_id == Person.id)
        .outerjoin(Workstation, Device.workstation_id == Workstation.id)
    )
    if req.location_id:
        stmt = stmt.where(Device.location_id == req.location_id)
    if req.person_id:
        stmt = stmt.where(Device.person_id == req.person_id)
    stmt = stmt.order_by(Device.inventory_number)
    result = await session.execute(stmt)
    rows = result.all()

    def _row(d, tn, tc, ln, pn, seat):
        cat = CATEGORY_LABELS.get(tc or "", tc or "")
        return (
            d.inventory_number or "",
            d.name or "",
            tn or "",
            cat,
            STATUS_LABELS.get(d.status, d.status),
            d.serial_number or "",
            d.model or "",
            d.manufacturer or "",
            ln or "",
            seat or "",
            pn or "",
            (d.commission_date.isoformat() if d.commission_date else ""),
            (d.purchase_date.isoformat() if d.purchase_date else ""),
            str(d.purchase_price) if d.purchase_price else "",
        )

    header = (
        "Инвентарный номер",
        "Наименование",
        "Тип устройства",
        "Категория типа",
        "Статус",
        "Серийный номер",
        "Модель",
        "Производитель",
        "Кабинет",
        "Рабочее место",
        "Ответственный",
        "Дата ввода",
        "Дата покупки",
        "Стоимость (₽)",
    )

    if req.format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(header)
        for r in rows:
            writer.writerow(_row(r[0], r[1], r[2], r[3], r[4], r[5]))
        return buf.getvalue().encode("utf-8-sig"), "text/csv", "devices-export.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Устройства"
    ws.append(header)
    for r in rows:
        ws.append(_row(r[0], r[1], r[2], r[3], r[4], r[5]))
    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "devices-export.xlsx",
    )


async def export_licenses(
    session: AsyncSession,
    req: ExportRequest,
) -> tuple[bytes, str, str]:
    """Экспорт лицензий в CSV или Excel."""
    stmt = select(License).order_by(License.name)
    result = await session.execute(stmt)
    licenses = list(result.scalars().all())

    def _row(lic: License) -> tuple[str, ...]:
        return (
            lic.name,
            str(lic.price),
            lic.expires_at.isoformat(),
            (lic.details or "").replace("\n", " ").replace("\r", "")[:500],
        )

    header = ("Название", "Стоимость (₽)", "Срок действия", "Подробности")

    if req.format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(header)
        for lic in licenses:
            writer.writerow(_row(lic))
        return buf.getvalue().encode("utf-8-sig"), "text/csv", "licenses-export.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Лицензии"
    ws.append(header)
    for lic in licenses:
        ws.append(_row(lic))
    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "licenses-export.xlsx",
    )


async def export_components(
    session: AsyncSession,
    req: ExportRequest,
) -> tuple[bytes, str, str]:
    """Экспорт комплектующих в CSV или Excel."""
    stmt = (
        select(
            Component,
            Device.inventory_number.label("host_inv"),
            Device.name.label("host_name"),
            Location.name.label("location_name"),
        )
        .outerjoin(ComputerComponent, ComputerComponent.component_id == Component.id)
        .outerjoin(Device, Device.id == ComputerComponent.computer_id)
        .outerjoin(Location, Device.location_id == Location.id)
        .order_by(Component.name)
    )
    if req.location_id:
        stmt = stmt.where(Device.location_id == req.location_id)
    result = await session.execute(stmt)
    rows = result.all()

    def _row(c: Component, host_inv, host_name, loc_name) -> tuple[str, ...]:
        type_label = COMPONENT_TYPE_LABELS.get(c.component_type, c.component_type)
        return (
            c.name,
            type_label,
            STATUS_LABELS.get(c.status, c.status),
            c.arrival_date.isoformat() if c.arrival_date else "",
            c.expiry_date.isoformat() if c.expiry_date else "",
            (c.notes or "").replace("\n", " ").replace("\r", "")[:300],
            host_inv or "",
            host_name or "",
            loc_name or "",
        )

    header = (
        "Название",
        "Тип комплектующей",
        "Статус",
        "Срок прибытия",
        "Срок годности",
        "Примечания",
        "Инв. № хоста",
        "Наименование хоста",
        "Кабинет хоста",
    )

    if req.format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(header)
        for r in rows:
            writer.writerow(_row(r[0], r[1], r[2], r[3]))
        return buf.getvalue().encode("utf-8-sig"), "text/csv", "components-export.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Комплектующие"
    ws.append(header)
    for r in rows:
        ws.append(_row(r[0], r[1], r[2], r[3]))
    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "components-export.xlsx",
    )


async def create_inventory_report(
    session: AsyncSession,
    req: InventoryRequest,
) -> InventoryResponse:
    """Акт инвентаризации."""
    loc = await session.get(Location, req.location_id)
    person = await session.get(Person, req.person_id)
    if loc is None:
        raise HTTPException(status_code=404, detail="Кабинет не найден")
    if person is None:
        raise HTTPException(status_code=404, detail="Ответственное лицо не найдено")

    stmt = (
        select(Device, DeviceType.name.label("type_name"), Workstation.seat_code.label("seat_code"))
        .outerjoin(DeviceType, Device.device_type_id == DeviceType.id)
        .outerjoin(Workstation, Device.workstation_id == Workstation.id)
        .where(
            and_(
                Device.location_id == req.location_id,
                Device.person_id == req.person_id,
                Device.status == "in_use",
            )
        )
        .order_by(Device.inventory_number)
    )
    result = await session.execute(stmt)
    rows = result.all()

    total = Decimal("0")
    items: list[InventoryDeviceItem] = []
    for d, type_name, seat_code in rows:
        if d.purchase_price:
            total += d.purchase_price
        items.append(
            InventoryDeviceItem(
                inventoryNumber=d.inventory_number,
                name=d.name,
                deviceTypeName=type_name,
                workstationSeat=seat_code,
                serialNumber=d.serial_number,
                status=STATUS_LABELS.get(d.status, d.status),
                purchasePrice=d.purchase_price,
            )
        )

    year = req.end_date.year
    doc_num = f"АКТ-{year}-{uuid.uuid4().hex[:6].upper()}"

    return InventoryResponse(
        id=uuid.uuid4(),
        documentNumber=doc_num,
        date=req.end_date,
        startDate=req.start_date,
        endDate=req.end_date,
        locationName=loc.name,
        personName=person.full_name,
        deviceCount=len(rows),
        totalPrice=total,
        devices=items,
    )
